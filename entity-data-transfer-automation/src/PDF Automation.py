import os
import PyPDF2
import pandas as pd
import re
import openpyxl
import textdistance


folder_path = r"C:\Users\AkimFitzgerald\Documents\Batch 5c Hits PDF"


def is_similar(name1, name2, threshold=0.85):
    return textdistance.jaro_winkler(name1, name2) >= threshold


# Loop over all PDF files in the folder
for filename in os.listdir(folder_path):
    if filename.endswith(".pdf"):  # Check if file is a PDF
        # Use regular expressions to extract text after the first dash ("-")
        match = re.search(r'^[^-]+-(.*)\.pdf', filename)
        if match:
            text_after_dash = match.group(1).title()  # Apply .title() method
            remaining_text = filename.replace("-" + match.group(1), "")
            if remaining_text.endswith(".pdf"):
                remaining_text = remaining_text[:-4]  # Remove ".pdf" extension
            new_filename = os.path.join(folder_path, remaining_text + ".pdf")
            os.rename(os.path.join(folder_path, filename), new_filename)
        else:
            print(filename, ": No match found.")


# Set the path to the folder containing the PDF files
path = r"C:\Users\AkimFitzgerald\Documents\Batch 5c Hits PDF"

# Initialize an empty list to store the parsed data
parsed_data = []

# Loop through each file in the folder
for filename in os.listdir(path):
    # Check if the file is a PDF
    if filename.endswith(".pdf"):
        # Open the PDF file
        with open(os.path.join(path, filename), "rb") as pdf_file:
            # Initialize a PDF reader object
            reader = PyPDF2.PdfReader(pdf_file)

            # Initialize an empty list to store the table data
            table_data = []

            # Loop through each page of the PDF file
            for page in reader.pages:
                # Extract the text from the page
                text = page.extract_text()

                # Split the text into rows using newline character
                rows = text.split("\n")

                # Loop through each row in the table
                for row in rows:
                    # Split the row into cells using tab character
                    cells = row.split("\t")

                    # Append the cells to the table data list
                    table_data.append(cells)

            # Create a pandas DataFrame from the table data and file name
            df = pd.DataFrame(table_data, columns=["Column 1"])
            df.insert(0, "File Name", filename)

            # Append the DataFrame to the parsed data list
            parsed_data.append(df)

# Concatenate all the DataFrames in the parsed data list
df = pd.concat(parsed_data, ignore_index=True)

# Use regular expressions to filter the rows that contain the specified strings
linkedin_mask = df["Column 1"].str.contains("linkedin\.com")
twitter_mask = df["Column 1"].str.contains("twitter\.com")
facebook_mask = df["Column 1"].str.contains("facebook\.com")

# Create a new DataFrame that includes only the rows that meet the mask conditions
filtered_df = df[linkedin_mask | twitter_mask | facebook_mask]


excel_folder_path = r"C:\Users\AkimFitzgerald\ORR Bulk Automation V2\Demo Folder that holds workbooks"

# Loop through each file in the Excel folder
for excel_filename in os.listdir(excel_folder_path):
    # Check if the file is an Excel file
    if excel_filename.endswith(".xlsx") or excel_filename.endswith(".xls"):
        
        # Load the Excel workbook
        wb = openpyxl.load_workbook(os.path.join(excel_folder_path, excel_filename))
        
        # Extract the sheet names
        sheet_names = [sheet.title.lower() for sheet in wb.worksheets]

        # Check if any PDF file has the same name as the sheet names in the Excel workbook (ignoring case)
        for pdf_filename in os.listdir(folder_path):
            if pdf_filename.endswith(".pdf"):
                for sheet_name in sheet_names:
                    if is_similar(pdf_filename[:-4].lower(), sheet_name):
                        print(f'Matched: {pdf_filename} with {excel_filename}')

                        # Read and store the PDF data only when a matching Excel workbook is found
                        with open(os.path.join(folder_path, pdf_filename), "rb") as pdf_file:
                            # Initialize a PDF reader object
                            reader = PyPDF2.PdfReader(pdf_file)

                            # Initialize an empty list to store the table data
                            table_data = []

                            # Loop through each page of the PDF file
                            for page in reader.pages:
                                # Extract the text from the page
                                text = page.extract_text()

                                # Split the text into rows using newline character
                                rows = text.split("\n")

                                # Loop through each row in the table
                                for row in rows:
                                    # Split the row into cells using tab character
                                    cells = row.split("\t")

                                    # Append the cells to the table data list
                                    table_data.append(cells)

                        # Create a pandas DataFrame from the table data and file name
                        pdf_df = pd.DataFrame(table_data, columns=["Column 1"])
                        pdf_df.insert(0, "File Name", pdf_filename)

                        # Use regular expressions to filter the rows that contain the specified strings
                        linkedin_mask = pdf_df["Column 1"].str.contains("linkedin\.com")
                        twitter_mask = pdf_df["Column 1"].str.contains("twitter\.com")
                        facebook_mask = pdf_df["Column 1"].str.contains("facebook\.com")

                        # Create a new DataFrame that includes only the rows that meet the mask conditions
                        filtered_pdf_df = pdf_df[linkedin_mask | twitter_mask | facebook_mask]

                        # Remove any existing sheet with the name "Social Media Data"
                        if "Social Media Data" in [sheet.title for sheet in wb.worksheets]:
                            wb.remove(wb["Social Media Data"])

                        # Create a new sheet with the name "Social Media Data"
                        ws = wb.create_sheet("Social Media Data")

                        # Write filtered_pdf_df data to the new sheet
                        for r, row in enumerate(filtered_pdf_df.values, start=1):
                            for c, value in enumerate(row, start=1):
                                ws.cell(row=r, column=c).value = value

                        # Save the updated Excel workbook
                        wb.save(os.path.join(excel_folder_path, excel_filename))






                
               
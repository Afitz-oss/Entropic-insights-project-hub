import os
import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Border, Side
import logging
import glob
from tqdm import tqdm

# Folder for text files 
folder_path = r"C:\Users\AkimFitzgerald\ORR Bulk Automation V2 Sample Stress Test (Not Sensitive Data)\Folder for text files"
# Folder for workbooks 
folder_path_excel = r"C:\Users\AkimFitzgerald\ORR Bulk Automation V2 Sample Stress Test (Not Sensitive Data)\Demo Folder that holds workbooks"
workbook_names = glob.glob(os.path.join(folder_path_excel, "*.xlsx"))
# Repeat the same workbook path
workbook_folder_path = r"C:\Users\AkimFitzgerald\ORR Bulk Automation V2 Sample Stress Test (Not Sensitive Data)\Demo Folder that holds workbooks"
# Repeat the same text path
text_file_folder_path = r"C:\Users\AkimFitzgerald\ORR Bulk Automation V2 Sample Stress Test (Not Sensitive Data)\Folder for text files"

# Sets font to bold for data entry
bold_font = Font(bold=True)

def extract_keys(pattern, folder_path):
    """
    Extracts the keys (file names without the extension) from all the text files in the specified folder
    and renames the files according to the extracted keys.

    Args:
        pattern (re.Pattern): A compiled regular expression pattern to match file names.
        folder_path (str): The path to the folder containing the text files.

    Returns:
        list: A list of keys (file names without the extension) from the text files in the folder.
    """

    # Initialize an empty list to store the keys.
    keys = []

    # Iterate over the files in the given folder.
    for file in os.listdir(folder_path):
        # Check if the file name matches the specified pattern.
        match = re.compile(r'(.+?)(-.*\.txt)').search(file)

        # If there's a match, append the key (file name without the extension) to the list.
        if match:
            key = match.group(1)
            keys.append(key)

            # Rename the file
            old_file_path = os.path.join(folder_path, file)
            new_file_path = os.path.join(folder_path, f'{key}.txt')

            # If the new file name already exists, remove the old file.
            if os.path.exists(new_file_path):
                os.remove(old_file_path)
            else:
                os.rename(old_file_path, new_file_path)

    # Return the list of keys.
    return keys

# Compile the regular expression pattern for matching file names.
pattern = re.compile(r'(.+?)(-.*\.txt)')

# Repeat the same text file path again.
folder_path = r"C:\Users\AkimFitzgerald\ORR Bulk Automation V2 Sample Stress Test (Not Sensitive Data)\Folder for text files"

# Call the `extract_keys` function with the compiled pattern and the folder path.
keys = extract_keys(pattern, folder_path)


# Print the extracted keys.
print(keys)


def txt_to_dict(directory, fields_data_dict, keys):
    """
    Reads text files, extracts specified fields from the file content, and stores the extracted information in a dictionary.

    Args:
        directory (str): The path to the directory containing the text files.
        fields_data_dict (dict): A dictionary with field keys and their corresponding start and end strings as values.
        keys (list): A list of keys (file names without the extension) from the text files in the folder.

    Returns:
        dict: A dictionary containing the extracted fields from the text files with keys as the file names without the extension.
    """

    # Initialize an empty dictionary to store the extracted data.
    text_file_dict = {}

    # Iterate over the keys.
    for key in keys:
        # Create the file path by joining the directory and the key with the '.txt' extension.
        file_path = os.path.join(directory, key + '.txt')

        # Check if the file path is valid and has a '.txt' extension.
        if os.path.isfile(file_path) and file_path.endswith('.txt'):
            # Open the text file for reading.
            with open(file_path, 'r') as f:
                # Read the file content.
                file_content = f.read()

                # Initialize an empty dictionary for the current key.
                text_file_dict[key] = {}

                # Iterate over the field keys and their corresponding start and end strings in the fields_data_dict.
                for field_key, (start, end) in fields_data_dict.items():
                    # Find the start index of the field.
                    start_index = file_content.find(start)

                    # If the start index is not found, set the field_key value to None and continue.
                    if start_index == -1:
                        text_file_dict[key][field_key] = None
                        continue

                    # Find the end index of the field.
                    end_index = file_content.find(end, start_index)

                    # If the end index is not found, set the field_key value to None and continue.
                    if end_index == -1:
                        text_file_dict[key][field_key] = None
                        continue

                    # Extract the field content and store it in the text_file_dict.
                    text_file_dict[key][field_key] = file_content[start_index + len(start): end_index].strip()

    # Return the dictionary containing the extracted data.
    return text_file_dict

fields_data_dict = {
    'Reference_ID': ('Reference ID:', 'Subject Information'),
    'Sponsor_Name': ('Name:', 'Date of Birth:'),
    'phone_number': ('Possible Phones Associated with Subject:', 'Indicators'),
    'Address': ('Address Summary ', 'Address Details'),
    'DOB': ('Date of Birth:', 'Gender:'),
    'Employer': ('Possible Employers', 'Address Summary'),
    'Associates': ('Possible Associates - Summary', 'Possible Associates - Details')}


# Chnge pth to se the same text file path that was previously used
results = txt_to_dict(r"C:\Users\AkimFitzgerald\ORR Bulk Automation V2 Sample Stress Test (Not Sensitive Data)\Folder for text files", fields_data_dict, keys)


excel_data_frames = []
for workbook_name in workbook_names:
    combined_excel_files = pd.read_excel(workbook_name)
    excel_data_frame = pd.DataFrame(combined_excel_files)
    excel_data_frame['Workbook_Name'] = workbook_name
    excel_data_frames.append(excel_data_frame)
df = pd.concat(excel_data_frames)

                     

def update_excel_with_text_data(workbook_folder_path, text_file_folder_path, fields_data_dict):
    """
    Updates Excel files in a given folder with data from text files in another given folder.

    Args:
    - workbook_folder_path (str): the path to the folder containing the Excel files to update.
    - text_file_folder_path (str): the path to the folder containing the text files with the data to update.
    - fields_data_dict (dict): a dictionary containing the keys to match between the text files and the Excel files, and
    the data to update in the Excel files.

    Returns:
    - None

    Raises:
    - Exception: if there is an error saving the updated Excel files.

    """
    logging.info('Starting update_excel_with_text_data')
    results = txt_to_dict(text_file_folder_path, fields_data_dict, keys)

    for workbook_file in os.listdir(workbook_folder_path):
        if workbook_file.endswith('.xlsx'):
            workbook_path = os.path.join(workbook_folder_path, workbook_file)
            workbook = openpyxl.load_workbook(workbook_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                found_match = False
                data_filled = True  # Initialize for each sheet
                for key in results:
                    if sheet_name.lower() == key.lower() or workbook_file.lower() == key.lower():
                        found_match = True
                        for field_key, field_value in results[key].items():
                            if field_key == 'Reference_ID':
                                sheet['A4'] = 'Uc Search name:'
                                sheet['A4'].font = bold_font
                                sheet['A5'] = field_value if field_value else 'No Results'
                                data_filled = data_filled and bool(field_value)
                            elif field_key == 'Sponsor_Name':
                                sheet['A6'] = 'Sponsors Search name:'
                                sheet['A6'].font = bold_font
                                sheet['A7'] = field_value
                            elif field_key == 'phone_number':
                                sheet['A10'] = 'Phone Number:'
                                sheet['A10'].font = bold_font
                                sheet['A11'] = field_value
                            elif field_key == 'Address':
                                sheet['A14'] = 'Address:'
                                sheet['A14'].font = bold_font
                                sheet['A15'] = field_value
                            elif field_key == 'DOB':
                                sheet['A18'] = 'Identification Information:'
                                sheet['A18'].font = bold_font
                                sheet['A19'] = field_value
                            elif field_key == 'Employer':
                                sheet['A21'] = 'Employer:'
                                sheet['A21'].font = bold_font
                                sheet['A22'] = field_value
                            elif field_key == 'Associates':
                                sheet['A25'] = 'Associates & Relatives:'
                                sheet['A25'].font = bold_font
                                sheet['A26'] = field_value
                        workbook.save(workbook_path)
                        break
                if not found_match:
                    sheet = workbook[sheet_name]
                    sheet['A4'] = 'Uc Search name:'
                    sheet['A4'].font = bold_font
                    sheet['A5'] = 'No Results'
                    sheet['A6'] = 'Sponsors Search name:'
                    sheet['A6'].font = bold_font
                    sheet['A7'] = 'No Results'
                    sheet['A10'] = 'Phone Number:'
                    sheet['A10'].font = bold_font
                    sheet['A11'] = 'No Results'
                    sheet['A14'] = 'Address:'
                    sheet['A14'].font = bold_font
                    sheet['A15'] = 'No Results'
                    sheet['A18'] = 'Identification Information:'
                    sheet['A18'].font = bold_font
                    sheet['A19'] = 'No Results'
                    sheet['A21'] = 'Employer:'
                    sheet['A21'].font = bold_font
                    sheet['A22'] = 'No Results'
                    sheet['A25'] = 'Associates & Relatives:'
                    sheet['A25'].font = bold_font
                    sheet['A26'] = 'No Results'
                    sheet.sheet_properties.tabColor = "FF0000"
                    logging.warning(f"No match found for key: {key} in workbook: {workbook_file}")
                if found_match:
                    if data_filled:
                        sheet.sheet_properties.tabColor = "00FF00"  # Green if all fields are filled
                    else:
                        sheet.sheet_properties.tabColor = "FFFF00"  # Yellow if partial data is filled
            try:
                workbook.save(workbook_path)
            except Exception as e:
                logging.error(f"Error saving workbook {workbook_path}: {e}")

    logging.info('Finished update_excel_with_text_data')



update_excel_with_text_data(workbook_folder_path, text_file_folder_path, fields_data_dict)







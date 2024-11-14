import datetime
from datetime import datetime
from time import strftime
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side

df1 = pd.read_excel(r"C:\Users\AkimFitzgerald\Stress_test_data.xlsx", sheet_name='RandomData')
columns = df1.columns


def create_workbook(item, columns, index, sheet_name, file_name):
    """
    Creates a new Excel workbook with a single sheet containing specified columns and data for a single item.

    Args:
        item (list): A list of values representing a single data item (row) to be added to the sheet.
        columns (list): A list of column names to be added to the sheet.
        index (int): The index of the row where the data for this item should be added.
        sheet_name (str): The name of the sheet where the data will be written.

    Returns:
        None
    """

    # Create font and fill styles
    bold_font = Font(bold=True)
    grey_fill = PatternFill(patternType='solid', fgColor='A0A0A0')
    dark_blue_fill = PatternFill(patternType='solid', fgColor='00008B')

    # Create border style
    top = Side(border_style='thick', color="A52A2A")
    bottom = Side(border_style='thick', color="A52A2A")
    border = Border(top=top, bottom=bottom, left=top, right=bottom)

    # Create a new workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    # Add the specified columns
    for i, column in enumerate(columns):
        worksheet.cell(row=1, column=i + 1).value = column

        # Set the column width to 35
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = 35

    # Add the data for this item
    for i, value in enumerate(item):
        # Check if the value is N/A, and replace it with "No Data"
        if pd.isna(value):
            value = "No Data"
        worksheet.cell(row=2, column=i + 1).value = value

    # Set the style for all cells in the worksheet
    for row in worksheet.rows:
        # Loop through all cells in the row
        for cell in row:
            # Set cell font to bold and fill to grey
            cell.font = bold_font
            cell.fill = grey_fill
            cell.border = border

    for row in worksheet.iter_rows(min_row=1, max_row=33):
        for cell in row:
            cell.border = openpyxl.styles.borders.Border(left=openpyxl.styles.Side(style='thin'),
                                                     right=openpyxl.styles.Side(style='thin'),
                                                     top=openpyxl.styles.Side(style='thin'),
                                                     bottom=openpyxl.styles.Side(style='thin'))


        # Save the workbook with the given file_name
    workbook.save(f"{file_name}.xlsx")



# Load the data from a sheet in an Excel file into a data frame
df = pd.read_excel(r"C:\Users\AkimFitzgerald\Stress_test_data.xlsx", sheet_name='RandomData')

# Iterate through each row of the DataFrame
for index, row in df.iterrows():
    # Check if 'Sponsor First Name' and 'Sponsor Last Name' are NaN
    if pd.isna(row['Sponsor First Name']) and pd.isna(row['Sponsor Last Name']):
        sponsor_name = "No Sponsor"
    else:
        sponsor_name = f"{row['Sponsor First Name']} {row['Sponsor Last Name']}"

    # Create a sheet name using the sponsor_name
    sheet_name = sponsor_name

    # To avoid any issues with applications that can't read sheet names longer than 31 characters,
    # truncate the sheet name to 31 characters
    sheet_name = sheet_name[:31]

    # Create a file name by concatenating 'UC Portal A#' and the sponsor_name
    file_name = f"{row['UC Portal A#']} {sponsor_name}"

    # To avoid any issues with applications that can't read file names longer than 31 characters,
    # truncate the file name to 31 characters
    file_name = file_name[:31]

    # Call the create_workbook function with the current row, columns, index, and the truncated sheet_name and file_name
    create_workbook(row, columns, index, sheet_name, file_name)



def process_row(df, index):
    
    # Extract the data from the row
    data1 = str(df['UC Portal A#'])
    data2 = str(df['Last Name'])
    data3 = str(df['First Name'])
    data4 = pd.to_datetime(df['DOB']).strftime("%m-%d-%Y")
    data5 = str(df['COB'])
    data6 = pd.to_datetime(df['Date of Discharge ']).strftime("%m-%d-%Y")
    data7 = str(df['Gender'])
    data8 = str(df['Sponsor First Name'])
    data9 = str(df['Sponsor Last Name'])
    data10 = pd.to_datetime(df['Sponsor DOB ']).strftime("%m-%d-%Y")
    data11 = str(df['Sponsor Relationship to UAC'])
    data12 = str(df['Sponsor Phone Number'])
    data13 = str(df['Last known address'])
    data14 = str(df['Address Flag in Portal? Address Used Multiple Times?'])
    data15 = str(df['Length of Residency at Known Address'])
    data16 = str(df['City '])
    data16 = str(df['City '])
    data17 = str(df['State'])
    data18 = str(df['Secondary Address'])
    data19 = str(df['Sponsor Email Address'])
    data20 = str(df['UAM School Name '])

    # Create a new workbook with the extracted data
    create_workbook([data1, data2, data3, data4, data5, data6, data7, data8, data9, data10, data11, data12, data13, data14, data15, data16, data17, data18, data19, data20], index)

df.apply(process_row, axis=1)